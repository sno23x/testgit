from flask import Blueprint, render_template, request, redirect, url_for, flash
from flask_login import login_required
from models import db, Customer, Sale
import openpyxl, io

customers_bp = Blueprint("customers", __name__)


def _next_cid():
    """ສ້າງ CID ຕໍ່ໄປ (CID001, CID002…)"""
    last = Customer.query.filter(Customer.cust_code.ilike("CID%"))\
        .order_by(Customer.cust_code.desc()).first()
    if last and last.cust_code:
        try:
            num = int(last.cust_code[3:]) + 1
        except ValueError:
            num = Customer.query.count() + 1
    else:
        num = Customer.query.count() + 1
    return f"CID{num:03d}"


@customers_bp.route("/next-cid")
@login_required
def next_cid():
    return {"cid": _next_cid()}


@customers_bp.route("/")
@login_required
def list_customers():
    q = request.args.get("q", "")
    debt_filter = request.args.get("debt", "all")   # all / has / none
    page = max(1, int(request.args.get("page", 1) or 1))
    query = Customer.query
    if q:
        query = query.filter(db.or_(
            Customer.name.ilike(f"%{q}%"),
            Customer.phone.ilike(f"%{q}%"),
            Customer.cust_code.ilike(f"%{q}%"),
        ))
    if debt_filter == "has":
        query = query.filter(Customer.total_debt > 0)
    elif debt_filter == "none":
        query = query.filter(Customer.total_debt <= 0)
    pagination = query.order_by(Customer.name).paginate(page=page, per_page=50, error_out=False)
    return render_template("customers/list.html",
                           customers=pagination.items, pagination=pagination,
                           q=q, debt_filter=debt_filter)


@customers_bp.route("/add", methods=["GET", "POST"])
@login_required
def add_customer():
    if request.method == "POST":
        cust_code = request.form.get("cust_code", "").strip()
        if not cust_code:
            cust_code = _next_cid()
        c = Customer(
            cust_code=cust_code,
            name=request.form.get("name", "").strip(),
            phone=request.form.get("phone", "").strip(),
            address=request.form.get("address", "").strip(),
        )
        db.session.add(c)
        db.session.commit()
        flash("ເພີ່ມລູກຄ້າສໍາເລັດ", "success")
        return redirect(url_for("customers.list_customers"))
    next_id = _next_cid()
    return render_template("customers/form.html", customer=None, next_cid=next_id)


@customers_bp.route("/<int:cid>/edit", methods=["GET", "POST"])
@login_required
def edit_customer(cid):
    c = Customer.query.get_or_404(cid)
    if request.method == "POST":
        c.cust_code = request.form.get("cust_code", c.cust_code).strip()
        c.name = request.form.get("name", c.name).strip()
        c.phone = request.form.get("phone", c.phone).strip()
        c.address = request.form.get("address", c.address).strip()
        db.session.commit()
        flash("ແກ້ໄຂສໍາເລັດ", "success")
        return redirect(url_for("customers.list_customers"))
    return render_template("customers/form.html", customer=c, next_cid=None)


@customers_bp.route("/import-excel", methods=["GET", "POST"])
@login_required
def import_excel():
    if request.method == "POST":
        f = request.files.get("excel_file")
        if not f:
            flash("ກະລຸນາເລືອກໄຟລ໌", "danger")
            return redirect(url_for("customers.import_excel"))
        try:
            wb = openpyxl.load_workbook(io.BytesIO(f.read()), read_only=True, data_only=True)
            ws = wb.active
        except Exception as e:
            flash(f"ໄຟລ໌ບໍ່ຖືກຕ້ອງ: {e}", "danger")
            return redirect(url_for("customers.import_excel"))

        # Auto-detect header row: look for ຊື່ ຫຼື NAME ຫຼື CUST
        header_row = None
        col_map = {}
        for ridx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
            cells = [str(c).strip().upper() if c else "" for c in row]
            if any(k in " ".join(cells) for k in ["ຊື່", "NAME", "CUST"]):
                header_row = ridx
                for cidx, cell in enumerate(cells):
                    if any(k in cell for k in ["CUST", "CID"]):
                        col_map["cust_code"] = cidx
                    elif any(k in cell for k in ["ຊື່", "NAME"]):
                        col_map["name"] = cidx
                    elif any(k in cell for k in ["ທີ່ຢູ່", "ADDRESS", "ADDR"]):
                        col_map["address"] = cidx
                    elif any(k in cell for k in ["ເບີ", "PHONE", "TEL", "MOBILE"]):
                        col_map["phone"] = cidx
                break

        if header_row is None or "name" not in col_map:
            flash("ຊອກບໍ່ພົບ header row (ຕ້ອງມີ column ຊື່ ຫຼື NAME)", "danger")
            return redirect(url_for("customers.import_excel"))

        added = updated = skipped = 0
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            name = str(row[col_map["name"]] or "").strip()
            if not name:
                skipped += 1
                continue
            cust_code = str(row[col_map.get("cust_code", -1)] or "").strip() if col_map.get("cust_code") is not None else ""
            phone = str(row[col_map.get("phone", -1)] or "").strip() if col_map.get("phone") is not None else ""
            address = str(row[col_map.get("address", -1)] or "").strip() if col_map.get("address") is not None else ""

            # Match by cust_code first, then by phone, then create new
            existing = None
            if cust_code:
                existing = Customer.query.filter_by(cust_code=cust_code).first()
            if not existing and phone:
                existing = Customer.query.filter_by(phone=phone).first()

            if existing:
                existing.name = name
                if phone: existing.phone = phone
                if address: existing.address = address
                if cust_code: existing.cust_code = cust_code
                updated += 1
            else:
                db.session.add(Customer(
                    cust_code=cust_code, name=name, phone=phone, address=address))
                added += 1

        db.session.commit()
        flash(f"ນໍາເຂົ້າສໍາເລັດ: ເພີ່ມ {added} | ອັບເດດ {updated} | ຂ້າມ {skipped}", "success")
        return redirect(url_for("customers.list_customers"))

    return render_template("customers/import.html")


@customers_bp.route("/<int:cid>")
@login_required
def customer_detail(cid):
    c = Customer.query.get_or_404(cid)
    sales = Sale.query.filter_by(customer_id=cid).order_by(Sale.created_at.desc()).all()
    return render_template("customers/detail.html", customer=c, sales=sales)
