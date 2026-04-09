import io, os, uuid
from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, current_app, send_file
from flask_login import login_required
from werkzeug.utils import secure_filename
from models import db, Product, Category, Setting, round_price

ALLOWED_EXT = {"png", "jpg", "jpeg", "gif", "webp"}

def _save_image(file_obj, old_filename=""):
    """Save uploaded image, return filename or old_filename if no new file."""
    if not file_obj or file_obj.filename == "":
        return old_filename
    ext = file_obj.filename.rsplit(".", 1)[-1].lower()
    if ext not in ALLOWED_EXT:
        return old_filename
    fname = f"{uuid.uuid4().hex}.{ext}"
    upload_dir = os.path.join(current_app.static_folder, "uploads", "products")
    os.makedirs(upload_dir, exist_ok=True)
    file_obj.save(os.path.join(upload_dir, fname))
    return fname


COST_THB_RATE = 700  # ອັດຕາແລກປ່ຽນຕົ້ນທຶນ: 1 ບາດ = 700 ກີບ


def _compute_sell_price(price_thb, sell_price_manual):
    """ຖ້າໃສ່ລາຄາ THB ຈະຄຳນວນ LAK ອັດຕະໂນມັດ, ຖ້າບໍ່ໃຊ້ ລາຄາ LAK ທີ່ກຳນົດ"""
    thb = float(price_thb or 0)
    if thb > 0:
        try:
            rate = float(Setting.get("thb_to_lak", "830"))
        except ValueError:
            rate = 830.0
        return round_price(thb * rate), thb
    return float(sell_price_manual or 0), None


def _compute_cost_price(cost_thb, cost_lak_manual):
    """ຄິດໄລ່ລາຄາຕົ້ນທຶນ: ຖ້າໃສ່ THB ໃຊ້ × 700, ຖ້າບໍ່ໃຊ້ LAK ທີ່ກຳນົດ"""
    thb = float(cost_thb or 0)
    if thb > 0:
        return round(thb * COST_THB_RATE)
    return float(cost_lak_manual or 0)

products_bp = Blueprint("products", __name__)


@products_bp.route("/")
@login_required
def list_products():
    q = request.args.get("q", "")
    cat_id = request.args.get("cat", "")
    page = max(1, int(request.args.get("page", 1) or 1))
    query = Product.query.filter_by(active=True)
    if q:
        query = query.filter(db.or_(Product.name.ilike(f"%{q}%"), Product.code.ilike(f"%{q}%")))
    if cat_id:
        query = query.filter_by(category_id=int(cat_id))
    pagination = query.order_by(Product.name).paginate(page=page, per_page=50, error_out=False)
    categories = Category.query.order_by(Category.name).all()
    return render_template("products/list.html",
                           products=pagination.items, pagination=pagination,
                           categories=categories, q=q, cat_id=cat_id)


@products_bp.route("/add", methods=["GET", "POST"])
@login_required
def add_product():
    categories = Category.query.order_by(Category.name).all()
    rate = Setting.get("thb_to_lak", "830")
    if request.method == "POST":
        code = request.form.get("code", "").strip()
        if not code:
            last = Product.query.order_by(Product.id.desc()).first()
            code = f"P{(last.id + 1 if last else 1):05d}"
        sell_price, price_thb = _compute_sell_price(
            request.form.get("price_thb"), request.form.get("sell_price")
        )
        cost_price = _compute_cost_price(
            request.form.get("cost_price_thb"), request.form.get("cost_price")
        )
        p = Product(
            code=code,
            name=request.form.get("name", "").strip(),
            unit=request.form.get("unit", "ອັນ").strip(),
            cost_price=cost_price,
            price_thb=price_thb,
            sell_price=sell_price,
            stock_qty=float(request.form.get("stock_qty", 0) or 0),
            category_id=int(request.form.get("category_id")) if request.form.get("category_id") else None,
            image=_save_image(request.files.get("image")),
        )
        db.session.add(p)
        db.session.commit()
        flash("ເພີ່ມສິນຄ້າສໍາເລັດ", "success")
        back_q = request.form.get("back_q", "")
        back_cat = request.form.get("back_cat", "")
        return redirect(url_for("products.list_products", q=back_q or None, cat=back_cat or None))
    back_q = request.args.get("back_q", "")
    back_cat = request.args.get("back_cat", "")
    return render_template("products/form.html", product=None, categories=categories, rate=rate,
                           back_q=back_q, back_cat=back_cat, cost_thb_rate=COST_THB_RATE)


@products_bp.route("/<int:pid>/edit", methods=["GET", "POST"])
@login_required
def edit_product(pid):
    p = Product.query.get_or_404(pid)
    categories = Category.query.order_by(Category.name).all()
    rate = Setting.get("thb_to_lak", "830")
    if request.method == "POST":
        p.code = request.form.get("code", p.code).strip()
        p.name = request.form.get("name", p.name).strip()
        p.unit = request.form.get("unit", p.unit).strip()
        p.cost_price = _compute_cost_price(
            request.form.get("cost_price_thb"), request.form.get("cost_price")
        )
        p.stock_qty = float(request.form.get("stock_qty", 0) or 0)
        p.category_id = int(request.form.get("category_id")) if request.form.get("category_id") else None
        sell_price, price_thb = _compute_sell_price(
            request.form.get("price_thb"), request.form.get("sell_price")
        )
        p.price_thb = price_thb
        p.sell_price = sell_price
        new_img = _save_image(request.files.get("image"), p.image or "")
        p.image = new_img
        db.session.commit()
        flash("ແກ້ໄຂສໍາເລັດ", "success")
        back_q = request.form.get("back_q", "")
        back_cat = request.form.get("back_cat", "")
        return redirect(url_for("products.list_products", q=back_q or None, cat=back_cat or None))
    back_q = request.args.get("back_q", "")
    back_cat = request.args.get("back_cat", "")
    return render_template("products/form.html", product=p, categories=categories, rate=rate,
                           back_q=back_q, back_cat=back_cat, cost_thb_rate=COST_THB_RATE)


@products_bp.route("/<int:pid>/delete", methods=["POST"])
@login_required
def delete_product(pid):
    p = Product.query.get_or_404(pid)
    p.active = False
    db.session.commit()
    flash("ລຶບສິນຄ້າສໍາເລັດ", "success")
    return redirect(url_for("products.list_products"))


# --- Categories ---
@products_bp.route("/categories")
@login_required
def list_categories():
    cats = Category.query.order_by(Category.name).all()
    return render_template("products/categories.html", categories=cats)


@products_bp.route("/categories/add", methods=["POST"])
@login_required
def add_category():
    name = request.form.get("name", "").strip()
    if name:
        db.session.add(Category(name=name))
        db.session.commit()
        flash("ເພີ່ມໝວດໝູ່ສໍາເລັດ", "success")
    return redirect(url_for("products.list_categories"))


@products_bp.route("/categories/<int:cid>/delete", methods=["POST"])
@login_required
def delete_category(cid):
    cat = Category.query.get_or_404(cid)
    db.session.delete(cat)
    db.session.commit()
    flash("ລຶບໝວດໝູ່ສໍາເລັດ", "success")
    return redirect(url_for("products.list_categories"))


# --- Bulk operations ---
@products_bp.route("/bulk-delete", methods=["POST"])
@login_required
def bulk_delete():
    data = request.get_json()
    ids = data.get("ids", [])
    if not ids:
        return jsonify({"ok": False, "error": "ບໍ່ມີ ID"})
    Product.query.filter(Product.id.in_(ids)).update({"active": False}, synchronize_session=False)
    db.session.commit()
    return jsonify({"ok": True})


@products_bp.route("/bulk-category", methods=["POST"])
@login_required
def bulk_category():
    data = request.get_json()
    ids = data.get("ids", [])
    cat_id = data.get("category_id")
    if not ids or not cat_id:
        return jsonify({"ok": False, "error": "ຂໍ້ມູນບໍ່ຄົບ"})
    Product.query.filter(Product.id.in_(ids)).update({"category_id": cat_id}, synchronize_session=False)
    db.session.commit()
    return jsonify({"ok": True})


# --- Live search API ---
@products_bp.route("/api/search")
@login_required
def api_search():
    q = request.args.get("q", "").strip()
    if not q:
        return jsonify([])
    results = Product.query.filter(
        Product.active == True,
        db.or_(Product.name.ilike(f"%{q}%"), Product.code.ilike(f"%{q}%"))
    ).order_by(Product.name).limit(20).all()
    return jsonify([{
        "id": p.id, "code": p.code, "name": p.name,
        "unit": p.unit, "sell_price": p.sell_price,
        "stock_qty": p.stock_qty,
    } for p in results])


# --- Export Excel ---
@products_bp.route("/export")
@login_required
def export_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    products = Product.query.filter_by(active=True).order_by(Product.name).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ສິນຄ້າ"

    headers = ["ລະຫັດ", "ຊື່ສິນຄ້າ", "ໝວດໝູ່", "ໜ່ວຍ",
               "ລາຄາຕົ້ນທຶນ (ກີບ)", "ລາຄາ (฿)", "ລາຄາຂາຍ (ກີບ)", "Stock"]
    col_widths = [12, 30, 15, 8, 18, 12, 18, 10]

    hdr_fill = PatternFill("solid", fgColor="4F81BD")
    hdr_font = Font(bold=True, color="FFFFFF")
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = w

    for row, p in enumerate(products, 2):
        ws.cell(row=row, column=1, value=p.code)
        ws.cell(row=row, column=2, value=p.name)
        ws.cell(row=row, column=3, value=p.category.name if p.category else "")
        ws.cell(row=row, column=4, value=p.unit)
        ws.cell(row=row, column=5, value=p.cost_price or 0)
        ws.cell(row=row, column=6, value=p.price_thb or "")
        ws.cell(row=row, column=7, value=p.sell_price)
        ws.cell(row=row, column=8, value=p.stock_qty)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="products.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
